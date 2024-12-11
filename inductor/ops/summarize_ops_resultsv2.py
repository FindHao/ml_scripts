#! /usr/bin/env python3
import argparse
import glob
import os
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

inductor_placeholder = "inductor"


def format_excel(writer, df, sheet_name, is_speedup=True):
    """Format Excel file with conditional formatting and column widths"""
    worksheet = writer.sheets[sheet_name]

    # Set column widths based on content
    for i, col in enumerate(df.columns):
        column_letter = openpyxl.utils.get_column_letter(i + 1)
        column_data = df[col].astype(str)
        max_length = max(max(len(str(x)) for x in column_data), len(col))
        worksheet.column_dimensions[column_letter].width = max_length * 1.2

    # Create cell formats
    red_fill = openpyxl.styles.PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    green_fill = openpyxl.styles.PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    number_format = openpyxl.styles.numbers.FORMAT_NUMBER_00

    # Apply number format to all numeric columns
    for col_num, col_name in enumerate(df.columns, 1):
        if col_name != "op_name":
            for row in range(2, len(df) + 2):  # Start from 2 to skip header
                cell = worksheet.cell(row=row, column=col_num)
                cell.number_format = number_format

    # Apply conditional formatting to comparison columns
    if is_speedup:
        compare_cols = [
            "p20_inductor_vs_liger",
            "p50_inductor_vs_liger",
            "p80_inductor_vs_liger",
        ]
    else:
        compare_cols = [
            "p20_inductor_vs_liger_mem",
            "p50_inductor_vs_liger_mem",
            "p80_inductor_vs_liger_mem",
        ]

    for col_name in compare_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                try:
                    value = float(cell.value)
                    if value >= 1.01:
                        cell.fill = green_fill
                    elif value <= 0.98:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    continue


def get_op_name(filename, liger_operators):
    """Extract op name from filename based on predefined operator list"""
    # Remove .csv extension and 'op_' prefix
    base_name = os.path.splitext(os.path.basename(filename))[0]
    if base_name.startswith("op_"):
        base_name = base_name[3:]

    # Find matching operator from the predefined list
    for op in liger_operators:
        if base_name.startswith(op):
            return op

    print("\n" + "!" * 80)
    print(f"WARNING: No matching operator found for {os.path.basename(filename)}")
    print("!" * 80 + "\n")
    return base_name


def process_folder_results(folder_path):
    """Process all CSV files in a folder and return speedup and memory summary data"""
    print(f"Processing folder: {os.path.basename(folder_path)}")
    csv_files = glob.glob(os.path.join(folder_path, "*.csv"))

    # Define known operators
    liger_operators = [
        "embedding",
        "rms_norm",
        "rope",
        "jsd",
        "fused_linear_jsd",
        "cross_entropy",
        "fused_linear_cross_entropy",
        "geglu",
        "kl_div",
        "swiglu",
        "layer_norm",
    ]

    speedup_results = []
    memory_results = []

    for csv_file in csv_files:
        print(f"Processing CSV file: {csv_file}")
        df = pd.read_csv(csv_file, sep=";")

        try:
            # Extract op name using predefined list
            op_name = get_op_name(csv_file, liger_operators)
            if op_name == "layer_norm":
                tmp_inductor_placeholder = "torch_compile"
            else:
                tmp_inductor_placeholder = inductor_placeholder
            # Find relevant columns
            liger_speedup_col = [
                col for col in df.columns if "liger" in col and "-speedup" in col
            ][0]
            inductor_speedup_col = [
                col
                for col in df.columns
                if tmp_inductor_placeholder in col and "-speedup" in col
            ][0]
            liger_mem_col = [
                col
                for col in df.columns
                if "liger" in col and "-mem_footprint_compression_ratio" in col
            ][0]
            inductor_mem_col = [
                col
                for col in df.columns
                if tmp_inductor_placeholder in col
                and "-mem_footprint_compression_ratio" in col
            ][0]

            # Convert metrics to numeric type
            liger_speedup = pd.to_numeric(df[liger_speedup_col], errors="coerce")
            inductor_speedup = pd.to_numeric(df[inductor_speedup_col], errors="coerce")
            liger_mem = pd.to_numeric(df[liger_mem_col], errors="coerce")
            inductor_mem = pd.to_numeric(df[inductor_mem_col], errors="coerce")

            # Skip files with insufficient valid data
            insufficient_data = {
                "liger_speedup": len(liger_speedup.dropna()),
                "inductor_speedup": len(inductor_speedup.dropna()),
                "liger_mem": len(liger_mem.dropna()),
                "inductor_mem": len(inductor_mem.dropna()),
            }
            if any(length == 0 for length in insufficient_data.values()):
                print(f"Skipping {csv_file} - insufficient valid data")
                for name, length in insufficient_data.items():
                    if length == 0:
                        print(f"    {name}: {length}")
                continue

            # Calculate ratios: inductor/liger
            inductor_vs_liger_speedup = inductor_speedup / liger_speedup
            inductor_vs_liger_mem = inductor_mem / liger_mem

            # Calculate percentiles for speedup metrics
            if len(liger_speedup.dropna()) > 0 and len(inductor_speedup.dropna()) > 0:
                speedup_metrics = {
                    "op_name": op_name,
                    "p20_inductor_vs_liger": np.percentile(
                        inductor_vs_liger_speedup.dropna(), 20
                    ),
                    "p50_inductor_vs_liger": np.percentile(
                        inductor_vs_liger_speedup.dropna(), 50
                    ),
                    "p80_inductor_vs_liger": np.percentile(
                        inductor_vs_liger_speedup.dropna(), 80
                    ),
                    "p20_liger_speedup": np.percentile(liger_speedup.dropna(), 20),
                    "p50_liger_speedup": np.percentile(liger_speedup.dropna(), 50),
                    "p80_liger_speedup": np.percentile(liger_speedup.dropna(), 80),
                    "p20_inductor_speedup": np.percentile(
                        inductor_speedup.dropna(), 20
                    ),
                    "p50_inductor_speedup": np.percentile(
                        inductor_speedup.dropna(), 50
                    ),
                    "p80_inductor_speedup": np.percentile(
                        inductor_speedup.dropna(), 80
                    ),
                }
                speedup_results.append(speedup_metrics)

            # Calculate percentiles for memory metrics
            if len(liger_mem.dropna()) > 0 and len(inductor_mem.dropna()) > 0:
                memory_metrics = {
                    "op_name": op_name,
                    "p20_inductor_vs_liger_mem": np.percentile(
                        inductor_vs_liger_mem.dropna(), 20
                    ),
                    "p50_inductor_vs_liger_mem": np.percentile(
                        inductor_vs_liger_mem.dropna(), 50
                    ),
                    "p80_inductor_vs_liger_mem": np.percentile(
                        inductor_vs_liger_mem.dropna(), 80
                    ),
                    "p20_liger_mem": np.percentile(liger_mem.dropna(), 20),
                    "p50_liger_mem": np.percentile(liger_mem.dropna(), 50),
                    "p80_liger_mem": np.percentile(liger_mem.dropna(), 80),
                    "p20_inductor_mem": np.percentile(inductor_mem.dropna(), 20),
                    "p50_inductor_mem": np.percentile(inductor_mem.dropna(), 50),
                    "p80_inductor_mem": np.percentile(inductor_mem.dropna(), 80),
                }
                memory_results.append(memory_metrics)

        except Exception as e:
            print("\n" + "=" * 80)
            print(f"ERROR processing {csv_file}:")
            print(f"    {str(e)}")
            print("=" * 80 + "\n")
            continue

    return pd.DataFrame(speedup_results), pd.DataFrame(memory_results)


def main():
    """Main function to process all folders and generate Excel reports"""
    # Add argument parser
    parser = argparse.ArgumentParser(
        description="Process benchmark results and generate Excel reports"
    )
    parser.add_argument(
        "-i",
        "--input-dir",
        type=str,
        required=True,
        help="Base path containing benchmark results",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=str,
        required=True,
        help="the output file path or name",
    )
    args = parser.parse_args()
    folder_path = args.input_dir
    if not args.output:
        output_file = folder_path + "/merged.xlsx"
    elif "/" not in args.output:
        output_file = folder_path + "/" + args.output + ".xlsx"
    elif args.output.endswith("/"):
        output_file = args.output + "merged.xlsx"
    else:
        raise ValueError(
            "Invalid output path: must be a directory or a filename without '/'"
        )

    speedup_df, memory_df = process_folder_results(folder_path)

    if os.path.exists(output_file):
        # Load existing workbook
        book = openpyxl.load_workbook(output_file)

        # Remove sheets if they already exist
        if "speedup_summary" in book.sheetnames:
            book.remove(book["speedup_summary"])
        if "memory_summary" in book.sheetnames:
            book.remove(book["memory_summary"])

        with pd.ExcelWriter(
            output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            # Write new sheets
            if not speedup_df.empty:
                speedup_df.to_excel(writer, sheet_name="speedup_summary", index=False)
                format_excel(writer, speedup_df, "speedup_summary", is_speedup=True)

            if not memory_df.empty:
                memory_df.to_excel(writer, sheet_name="memory_summary", index=False)
                format_excel(writer, memory_df, "memory_summary", is_speedup=False)

            # Get the workbook from the writer
            workbook = writer.book

            # Calculate current positions
            current_position_speedup = workbook.sheetnames.index("speedup_summary")

            # Move speedup_summary to the beginning
            workbook.move_sheet("speedup_summary", offset=-current_position_speedup)

            # Move memory_summary right after speedup_summary
            current_position_memory = workbook.sheetnames.index("memory_summary")
            target_position = 1  # Right after speedup_summary
            offset = target_position - current_position_memory
            workbook.move_sheet("memory_summary", offset=offset)

    else:
        # Create new file if it doesn't exist
        with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
            if not speedup_df.empty:
                speedup_df.to_excel(writer, sheet_name="speedup_summary", index=False)
                format_excel(writer, speedup_df, "speedup_summary", is_speedup=True)

            if not memory_df.empty:
                memory_df.to_excel(writer, sheet_name="memory_summary", index=False)
                format_excel(writer, memory_df, "memory_summary", is_speedup=False)

    print(f"Results saved to {output_file}")


if __name__ == "__main__":
    main()
