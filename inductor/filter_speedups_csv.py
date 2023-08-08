import pandas as pd
import os
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
def get_file_paths(file):
    with open(file, 'r') as f:
        return [line.strip() for line in f]

def process_accuracy_file(file):
    df = pd.read_csv(file)
    return df[['name', 'accuracy']]

def process_performance_file(file):
    df = pd.read_csv(file)
    return df[['name', 'speedup', 'abs_latency']]

def process_files(file_paths, output_file):
    df_accuracy_dict = {}
    df_performance_dict = {}
    df_single_stream_dict = {}
    file_names_dict = {}  # Dictionary to store CSV file names

    for file in file_paths:
        filename = os.path.basename(file)
        if 'compilation_metrics' in filename:
            continue
        if 'accuracy' in filename:
            collection = filename.split('_')[2]
            df = process_accuracy_file(file)
            if collection not in df_accuracy_dict:
                df_accuracy_dict[collection] = df
            else:
                df_accuracy_dict[collection] = df_accuracy_dict[collection].append(df)
            # Add file name to dictionary
            file_names_dict[collection] = file_names_dict.get(collection, []) + [filename]
        elif 'performance' in filename and 'single_stream' not in filename:
            collection = filename.split('_')[2]
            df = process_performance_file(file)
            if collection not in df_performance_dict:
                df_performance_dict[collection] = df
            else:
                df_performance_dict[collection] = df_performance_dict[collection].append(df)
            file_names_dict[collection] = file_names_dict.get(collection, []) + [filename]
        elif 'single_stream' in filename:
            collection = filename.split('_')[2]
            df = process_performance_file(file)
            if collection not in df_single_stream_dict:
                df_single_stream_dict[collection] = df
            else:
                df_single_stream_dict[collection] = df_single_stream_dict[collection].append(df)
            file_names_dict[collection] = file_names_dict.get(collection, []) + [filename]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        if df_accuracy_dict and df_performance_dict and df_single_stream_dict:
            for collection in df_accuracy_dict:
                if collection in df_performance_dict and collection in df_single_stream_dict:
                    df_accuracy = df_accuracy_dict[collection]
                    df_performance = df_performance_dict[collection]
                    df_single_stream = df_single_stream_dict[collection]

                    df_merge = pd.merge(df_accuracy, df_performance, on='name', how='outer')
                    df_merge.rename(columns={'accuracy': 'accuracy', 'speedup': 'speedup', 'abs_latency': 'abs_latency'}, inplace=True)

                    df_single_stream.rename(columns={'speedup': 'original speedup'}, inplace=True)
                    df_merge = pd.merge(df_merge, df_single_stream[['name', 'original speedup']], on='name', how='outer')

                    df_merge['changes'] = df_merge['speedup'] - df_merge['original speedup']
                    df_merge['change speedup'] = df_merge['changes'] / df_merge['original speedup']
                    # Round the specified columns to 3 decimal places
                    df_merge[['speedup', 'abs_latency', 'original speedup', 'changes', 'change speedup']] = df_merge[['speedup', 'abs_latency', 'original speedup', 'changes', 'change speedup']].round(3)
                    # Reorder the columns
                    df_merge = df_merge.reindex(['name', 'accuracy', 'abs_latency', 'original speedup', 'speedup', 'changes', 'change speedup'], axis=1)
                    # Sort the dataframe by 'accuracy' and 'changes' columns
                    df_merge = df_merge.sort_values(by=['accuracy', 'change speedup'])
                    df_merge.to_excel(writer, sheet_name=collection, index=False)
                    # Get the worksheet of the current collection
                    ws = writer.sheets[collection]
                    # Insert file names at the top of the sheet
                    # Inserting a row at the beginning shifts existing rows down
                    ws.insert_rows(1)
                    ws.cell(row=1, column=1, value="Original CSV Files: " + ', '.join(file_names_dict[collection]))
        else:
            print("No data found for processing.")

from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import openpyxl
def format_excel_file(filename):
    # Open the workbook
    wb = openpyxl.load_workbook(filename)

    # Define the fills
    red_fill = PatternFill(start_color="F38A96", end_color="F38A96", fill_type="solid")
    green_fill = PatternFill(start_color="10BE07", end_color="10BE07", fill_type="solid")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Add conditional formatting rule for the 'changes' column
        red_rule = Rule(type="expression", dxf=DifferentialStyle(fill=red_fill))
        red_rule.formula = ["$G3<-0.02"]
        sheet.conditional_formatting.add("G3:G1048576", red_rule)

        green_rule = Rule(type="expression", dxf=DifferentialStyle(fill=green_fill))
        green_rule.formula = ["$G3>0.02"]
        sheet.conditional_formatting.add("G3:G1048576", green_rule)

        # For each column in the worksheet
        for column in sheet.iter_cols(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            # Adjust the width of the column to the maximum length found (plus a little extra for padding)
            sheet.column_dimensions[column_letter].width = max_length + 2
    wb.save(filename)


if __name__ == '__main__':
    # get current date and time
    import datetime
    now = datetime.datetime.now()
    now_str = now.strftime("%Y%m%d_%H%M")
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', type=str, default="/tmp/yhao/input.txt")
    parser.add_argument('--output', type=str, default=f"/home/users/yhao24/b/tmp/profile/output_{now_str}.xlsx")
    args = parser.parse_args()
    file_paths = get_file_paths(args.input)
    process_files(file_paths, args.output)
    format_excel_file(args.output)
    print(f"The output file is {args.output}")
