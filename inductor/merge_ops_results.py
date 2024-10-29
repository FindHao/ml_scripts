import pandas as pd
import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
import argparse


def parse_args():
    parser = argparse.ArgumentParser(
        description="Merge CSV files into a single Excel file."
    )
    parser.add_argument(
        "--input",
        type=str,
        required=True,
        help="Input folder path containing CSV files",
    )
    parser.add_argument(
        "--output",
        type=str,
        required=True,
        help="Output path for the merged Excel file",
    )
    return parser.parse_args()


def create_chart(
    worksheet,
    start_row,
    end_row,
    cols,
    title,
    chart_position,
    categories_col,
    is_speedup,
):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.height = 10
    chart.width = 20

    # Add data series only for specified columns
    for col in cols:
        data = Reference(worksheet, min_col=col, min_row=1, max_row=end_row)
        series = Series(data, title_from_data=True)

        # Get original name and remove suffix
        series_name = worksheet.cell(row=1, column=col).value
        if is_speedup:
            series_name = series_name.replace("-speedup", "")
        else:
            series_name = series_name.replace("-mem_footprint", "")

        # Create proper SeriesLabel
        series.tx = SeriesLabel(v=series_name)
        chart.series.append(series)

    # Set categories (x-axis)
    cats = Reference(worksheet, min_col=categories_col, min_row=2, max_row=end_row)
    chart.set_categories(cats)

    # Add data labels
    for series in chart.series:
        series.data_labels = DataLabelList()
        series.data_labels.number_format = "0.00"

    # Add axis titles
    chart.x_axis.title = worksheet.cell(row=1, column=categories_col).value
    if is_speedup:
        chart.y_axis.title = "Speedup"
    else:
        chart.y_axis.title = "Memory Footprint"

    worksheet.add_chart(chart, chart_position)


def main():
    args = parse_args()

    # Get folder path and output path from arguments
    folder_path = args.input
    output_path = args.output

    # Ensure folder path ends with '/'
    if not folder_path.endswith("/"):
        folder_path += "/"

    # Specify file pattern for CSV files
    file_pattern = "*.csv"

    # Read all CSV files
    csv_files = glob.glob(folder_path + file_pattern)

    # Create an Excel writer object
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for file in csv_files:
            # Read each CSV file into a DataFrame with the correct delimiter
            df = pd.read_csv(file, delimiter=";")

            # Round numeric columns to 2 decimal places
            df = df.round(2)

            # Extract the middle part of the filename as the sheet name
            filename_parts = file.split("/")[-1].split("_")
            sheet_name = "_".join(filename_parts[1:-1])

            # Write each DataFrame to a separate sheet with the extracted name
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get the worksheet
            worksheet = writer.sheets[sheet_name]

            # Debug: print column names
            print("All columns:", df.columns.tolist())

            # More precise column selection
            speedup_cols = [
                i + 1
                for i, col in enumerate(df.columns)
                if col.strip().endswith("speedup")
            ]
            mem_cols = [
                i + 1
                for i, col in enumerate(df.columns)
                if col.strip().endswith("mem_footprint")
            ]

            # Debug: print selected columns
            print(
                "Selected speedup columns:", [df.columns[i - 1] for i in speedup_cols]
            )
            print("Selected memory columns:", [df.columns[i - 1] for i in mem_cols])

            # Find x_val column
            x_val_col = 1

            if speedup_cols:
                create_chart(
                    worksheet=worksheet,
                    start_row=1,
                    end_row=len(df) + 1,
                    cols=speedup_cols,
                    title=f"{sheet_name} Speedup",
                    chart_position="A10",
                    categories_col=x_val_col,
                    is_speedup=True,
                )

            if mem_cols:
                create_chart(
                    worksheet=worksheet,
                    start_row=1,
                    end_row=len(df) + 1,
                    cols=mem_cols,
                    title=f"{sheet_name} Memory Footprint",
                    chart_position="P10",
                    categories_col=x_val_col,
                    is_speedup=False,
                )

    print(f"All CSV files merged successfully into {output_path}")


if __name__ == "__main__":
    main()
