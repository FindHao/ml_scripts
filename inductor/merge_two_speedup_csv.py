import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import argparse
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule

green_fill = PatternFill(start_color="10be07", end_color="10be07", fill_type="solid")
red_fill = PatternFill(start_color="f38a96", end_color="f38a96", fill_type="solid")
def apply_coloring(ws):
    # Define fill colors

    # Create rules
    green_rule = CellIsRule(operator="greaterThan", formula=['0.02'], fill=green_fill)
    red_rule = CellIsRule(operator="lessThan", formula=['-0.02'], fill=red_fill)

    # Apply rules to desired columns
    for col in ['G', 'M']:
        ws.conditional_formatting.add(f"{col}3:{col}{ws.max_row}", green_rule)
        ws.conditional_formatting.add(f"{col}3:{col}{ws.max_row}", red_rule)

# Example usage:
# apply_coloring(your_worksheet)

def merge_sheets_based_on_model(ws1, ws2, ws_out):
    # Copy headers first from the first worksheet
    for cell in ws1[1]:
        ws_out.cell(row=1, column=cell.col_idx, value=cell.value)

    # Insert new row for "no bypass" and "bypass" and merge as necessary
    ws_out.insert_rows(idx=2)
    ws_out['B2'] = "no bypass"
    ws_out['H2'] = "bypass"
    ws_out.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    ws_out.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)
    # Center align the merged cells
    alignment = Alignment(horizontal="center", vertical="center")
    ws_out['B2'].alignment = alignment
    ws_out['H2'].alignment = alignment
    ws1_dict = {row[0]: row for row in ws1.iter_rows(min_row=2, values_only=True) if row[0]}
    ws2_dict = {row[0]: row for row in ws2.iter_rows(min_row=2, values_only=True) if row[0]}

    for idx, model_name in enumerate(ws1_dict.keys(), start=3):  # starting from 3 because of the extra row we added
        # Writing values from the first sheet
        for col_idx, cell_value in enumerate(ws1_dict[model_name], start=1):
            ws_out.cell(row=idx, column=col_idx, value=cell_value)

        # Writing values from the second sheet
        if model_name in ws2_dict:
            for col_idx, cell_value in enumerate(ws2_dict[model_name][1:], start=ws1.max_column + 1):
                ws_out.cell(row=idx, column=col_idx, value=cell_value)


def apply_additional_column_and_coloring(ws):
    # Calculate the last column's letter
    last_col_letter = get_column_letter(ws.max_column + 1)
    ws[f"{last_col_letter}3"].value = "original speedup diff"
    # Populate the new column with the formula
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), 4):
        colD_cell = row[3]  # 0-based index, so 3 corresponds to column D
        colJ_cell = row[9]  # 0-based index, so 9 corresponds to column J
        
        if colD_cell.value is not None and colJ_cell.value is not None:
            new_cell = ws[f"{last_col_letter}{i}"]
            new_cell.value = f"={colD_cell.coordinate}/{colJ_cell.coordinate} - 1"
            new_cell.number_format = "0.00"  # Two decimal places

     # Create rules
    positive_red_rule = FormulaRule(formula=[f"{last_col_letter}4>0.02"], fill=red_fill)
    negative_red_rule = FormulaRule(formula=[f"{last_col_letter}4<-0.02"], fill=red_fill)

    # Apply rules to the new column
    ws.conditional_formatting.add(f"{last_col_letter}4:{last_col_letter}{ws.max_row}", positive_red_rule)
    ws.conditional_formatting.add(f"{last_col_letter}4:{last_col_letter}{ws.max_row}", negative_red_rule)




def merge_excel_files(file1, file2, output_name="merged.xlsx"):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)  # remove default sheet

    for sheet_name in wb1.sheetnames:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        ws_out = output_wb.create_sheet(sheet_name)
        
        # Merge based on model name
        merge_sheets_based_on_model(ws1, ws2, ws_out)
        apply_coloring(ws_out)
        apply_additional_column_and_coloring(ws_out)
    output_wb.save(output_name)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Merge two Excel files.')
    parser.add_argument('--input', required=True, help='Path to the first input Excel file.')
    parser.add_argument('--input-bypass', required=True, help='Path to the second input Excel file.')
    parser.add_argument('--output', default='merged.xlsx', help='Path to the file.')
    args = parser.parse_args()
    merge_excel_files(args.input, args.input_bypass, args.output)

